#!/usr/bin/env python3.11
"""
validate_configs.py — Config coverage validation report

Reads all 642 unique configs from wis_unique_configs, cross-references .in/.com
CSVs and PMS Description Cleaned sheets, queries Jira SQLite for ticket mentions,
checks wiki node connections, then writes a color-coded Excel report.

Output: raw/config_validation_report.xlsx
"""

import csv
import os
import re
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent.parent
RAW = BASE / "raw" / "modules"
WIKI = BASE / "wiki"
JIRA_DB = BASE / "raw" / "jira" / "tickets.sqlite"
OUTPUT = BASE / "raw" / "config_validation_report.xlsx"

UNIQUE_CONFIGS_CSV = RAW / "pms-configs-in" / "wis_unique_configs" / "Unique Configs.csv"

IN_DIR = RAW / "pms-configs-in" / "All WIS CONFIGS"
COM_DIR = RAW / "pms-configs-com" / "wis_service_configs"
DESC_DIR = RAW / "_root" / "Copy of Workplace_PMS Description (Cleaned)"

# ── Service → module slug mapping ─────────────────────────────────────────
SERVICE_TO_MODULE = {
    # unique configs Service(s) column values (varied naming)
    "MEETING ROOMS":                        "meeting-rooms",
    "Visitor Management Service":           "visitor-management",
    "VISITOR":                              "visitor-management",
    "VMS":                                  "visitor-management",
    "BOOKING-RULE-ENGINE":                  "booking-rule-engine",
    "Booking Rule Engine":                  "booking-rule-engine",
    "WIS-SEAT-BOOKING":                     "desk-management",
    "WIS Seat Booking":                     "desk-management",
    "GUARD_APP":                            "guard-app-kiosks",
    "Guard App":                            "guard-app-kiosks",
    "EMPLOYEE-EXPERIENCE-COMMON-CONFIG":    "employee-experience",
    "EMP_EXP_INTERNAL_CONFIGURATIONS":      "employee-experience",
    "EMPLOYEE_EXPERIENCE_COMMON_CONFIG":    "employee-experience",
    "EMP_EXPERIENCE_EMAIL_CONFIG":          "employee-experience",
    "Employee Experience":                  "employee-experience",
    "EMP_EXP":                              "employee-experience",
    "PROJECT-MANAGEMENT-SERVICE":           "pms",
    "PMS":                                  "pms",
    "APP_SERVER_CONFIG":                    "mobile-app",
    "MOBILE_APP_SERVER":                    "mobile-app",
    "Mobile App Server":                    "mobile-app",
    "AppServerConfig":                      "mobile-app",
}

# .in sheet name → slug
IN_SHEET_SLUG = {
    "1. PMS.csv":                    "pms",
    "2. Visitor Mgmt.csv":           "visitor-management",
    "3. Meeting Rooms.csv":          "meeting-rooms",
    "4. Booking Rule Engine.csv":    "booking-rule-engine",
    "5. WIS Seat Booking.csv":       "desk-management",
    "6. Guard App.csv":              "guard-app-kiosks",
    "7. Email Emp Experience.csv":   "employee-experience",
    "8. Emp Exp Internal Config.csv":"employee-experience",
    "9. Emp Exp Common Config.csv":  "employee-experience",
}

# .com sheet name → slug
COM_SHEET_SLUG = {
    "1. PMS.csv":                    "pms",
    "2. VMS.csv":                    "visitor-management",
    "3. Meeting Rooms.csv":          "meeting-rooms",
    "4. Booking Rule Engine.csv":    "booking-rule-engine",
    "5. WIS Seat Booking.csv":       "desk-management",
    "6. Guard App.csv":              "guard-app-kiosks",
    "7. Email Emp Experience.csv":   "employee-experience",
    "8. Emp Exp Internal Config.csv":"employee-experience",
    "9. Emp Exp Common Config.csv":  "employee-experience",
}

# desc sheet name → property name col, description col
DESC_SLUG = {
    "ProjectMgmtService.csv":  "pms",
    "Visitor.csv":             "visitor-management",
    "MeetingRooms.csv":        "meeting-rooms",
    "BookingRuleEngine.csv":   "booking-rule-engine",
    "WisSeatBooking.csv":      "desk-management",
    "GuardApp.csv":            "guard-app-kiosks",
    "EmpExpInternal.csv":      "employee-experience",
    "EmpExpCommon.csv":        "employee-experience",
    "MobileAppServer.csv":     "mobile-app",
    "AppServerConfig.csv":     "mobile-app",
}

# ── Colors ─────────────────────────────────────────────────────────────────
RED    = PatternFill("solid", fgColor="FFCDD2")   # no description
YELLOW = PatternFill("solid", fgColor="FFF9C4")   # partial coverage
GREEN  = PatternFill("solid", fgColor="C8E6C9")   # good coverage
GRAY   = PatternFill("solid", fgColor="F5F5F5")   # header
BOLD   = Font(bold=True)
THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── Loaders ────────────────────────────────────────────────────────────────

def load_unique_configs() -> list[dict]:
    """Returns list of {name, description, services, service_count}."""
    configs = []
    with open(UNIQUE_CONFIGS_CSV, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        in_data = False
        for row in reader:
            if not row or not row[0]:
                continue
            if row[0] == "Property Name":
                in_data = True
                continue
            if not in_data:
                continue
            name = row[0].strip()
            desc = row[1].strip() if len(row) > 1 else ""
            services = row[2].strip() if len(row) > 2 else ""
            svc_count = row[3].strip() if len(row) > 3 else "1"
            if name:
                configs.append({
                    "name": name,
                    "desc_unique": desc,
                    "services": services,
                    "service_count": svc_count,
                })
    return configs


def load_in_configs() -> dict[str, dict]:
    """Returns {prop_name: {slug, description}}."""
    result = {}
    for fname, slug in IN_SHEET_SLUG.items():
        path = IN_DIR / fname
        if not path.exists():
            continue
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            in_data = False
            for row in reader:
                if not row or not row[0]:
                    continue
                if row[0] == "Property Name":
                    in_data = True
                    continue
                if not in_data:
                    continue
                name = row[0].strip()
                desc = row[1].strip() if len(row) > 1 else ""
                if name:
                    # keep first occurrence (earlier service has priority)
                    if name not in result:
                        result[name] = {"slug": slug, "description": desc}
    return result


def load_com_configs() -> dict[str, dict]:
    """Returns {prop_name: {slug, data_type, description}}."""
    result = {}
    for fname, slug in COM_SHEET_SLUG.items():
        path = COM_DIR / fname
        if not path.exists():
            continue
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header_seen = False
            for row in reader:
                if not row or not row[0]:
                    continue
                if row[0] == "Property Name":
                    header_seen = True
                    continue
                if not header_seen:
                    continue
                name = row[0].strip()
                data_type = row[1].strip() if len(row) > 1 else ""
                desc = row[2].strip() if len(row) > 2 else ""
                if name:
                    if name not in result:
                        result[name] = {"slug": slug, "data_type": data_type, "description": desc}
    return result


def load_desc_cleaned() -> dict[str, str]:
    """Returns {prop_name: description} from PMS Description Cleaned sheets."""
    result = {}
    for fname in DESC_SLUG:
        path = DESC_DIR / fname
        if not path.exists():
            continue
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header_seen = False
            for row in reader:
                if not row or not row[0]:
                    continue
                if row[0] == "Property Name":
                    header_seen = True
                    continue
                if not header_seen:
                    continue
                name = row[0].strip()
                desc = row[1].strip() if len(row) > 1 else ""
                if name and name not in result:
                    result[name] = desc
    return result


def load_jira_mentions(prop_names: list[str]) -> dict[str, int]:
    """Returns {prop_name: ticket_count} using SQLite LIKE search."""
    if not JIRA_DB.exists():
        return {}
    conn = sqlite3.connect(str(JIRA_DB))
    cursor = conn.cursor()
    result = {}
    for name in prop_names:
        # Use parameterized LIKE; escape % and _ in name
        safe = name.replace("%", r"\%").replace("_", r"\_")
        pattern = f"%{safe}%"
        cursor.execute(
            """
            SELECT COUNT(*) FROM tickets
            WHERE summary LIKE ? ESCAPE '\\'
               OR description_text LIKE ? ESCAPE '\\'
               OR comments_text LIKE ? ESCAPE '\\'
            """,
            (pattern, pattern, pattern),
        )
        count = cursor.fetchone()[0]
        result[name] = count
    conn.close()
    return result


def get_wiki_nodes(prop_name: str, service_slug: str) -> dict:
    """Check which wiki nodes reference or relate to this config."""
    nodes = {
        "config_page": False,
        "module_page": False,
        "cross_module_pages": 0,
    }
    # Config page for the service
    if service_slug:
        cfg_path = WIKI / "configs" / f"{service_slug}.md"
        nodes["config_page"] = cfg_path.exists()
        mod_path = WIKI / "modules" / f"{service_slug}.md"
        nodes["module_page"] = mod_path.exists()

    # Cross-module pages mentioning this service
    cm_dir = WIKI / "cross-module"
    if cm_dir.exists() and service_slug:
        count = sum(
            1 for p in cm_dir.iterdir()
            if p.suffix == ".md" and service_slug in p.stem
        )
        nodes["cross_module_pages"] = count

    return nodes


def resolve_module(services_str: str) -> str:
    """Map the Services column value to a wiki module slug."""
    if not services_str:
        return ""
    # Handle multi-service (rare, but service_count > 1 cases)
    parts = [s.strip() for s in re.split(r"[,;|/]", services_str)]
    for part in parts:
        if part in SERVICE_TO_MODULE:
            return SERVICE_TO_MODULE[part]
    # fuzzy fallback
    low = services_str.lower()
    for key, val in SERVICE_TO_MODULE.items():
        if key.lower() in low:
            return val
    return services_str  # return raw if no mapping


def best_description(
    name: str,
    unique_desc: str,
    in_configs: dict,
    com_configs: dict,
    desc_cleaned: dict,
) -> tuple[str, str]:
    """Returns (description, source) using fallback chain."""
    # 1. .com
    if name in com_configs and com_configs[name]["description"]:
        d = com_configs[name]["description"]
        if d and d not in ("-", "—", "N/A"):
            return d, ".com"
    # 2. .in
    if name in in_configs and in_configs[name]["description"]:
        d = in_configs[name]["description"]
        if d and d not in ("-", "—", "N/A"):
            return d, ".in"
    # 3. unique configs master list
    if unique_desc:
        return unique_desc, "unique_configs"
    # 4. PMS Description Cleaned
    if name in desc_cleaned and desc_cleaned[name]:
        return desc_cleaned[name], "pms_desc_cleaned"
    return "", "none"


# ── Report builder ─────────────────────────────────────────────────────────

def build_report():
    print("Loading unique configs...")
    configs = load_unique_configs()
    print(f"  {len(configs)} configs loaded")

    print("Loading .in configs...")
    in_configs = load_in_configs()
    print(f"  {len(in_configs)} .in properties")

    print("Loading .com configs...")
    com_configs = load_com_configs()
    print(f"  {len(com_configs)} .com properties")

    print("Loading PMS Description Cleaned...")
    desc_cleaned = load_desc_cleaned()
    print(f"  {len(desc_cleaned)} descriptions")

    print("Querying Jira SQLite for ticket mentions (this may take a moment)...")
    prop_names = [c["name"] for c in configs]
    jira_counts = load_jira_mentions(prop_names)
    print("  Done.")

    print("Checking wiki nodes...")

    # ── Build rows ──────────────────────────────────────────────────────────
    rows = []
    for c in configs:
        name = c["name"]
        module = resolve_module(c["services"])

        in_present  = "✅" if name in in_configs else "—"
        com_present = "✅" if name in com_configs else "—"
        data_type   = com_configs[name]["data_type"] if name in com_configs else ""

        desc, desc_source = best_description(
            name, c["desc_unique"], in_configs, com_configs, desc_cleaned
        )

        jira_count = jira_counts.get(name, 0)
        wiki = get_wiki_nodes(name, module)

        # Connected node count: config page + module page + cross-module pages
        connected_nodes = (
            (1 if wiki["config_page"] else 0)
            + (1 if wiki["module_page"] else 0)
            + wiki["cross_module_pages"]
        )

        # ── Status logic ────────────────────────────────────────────────────
        has_desc = bool(desc)
        on_in    = (in_present == "✅")
        on_com   = (com_present == "✅")

        if not has_desc:
            status = "🔴 Flagged"
        elif not on_in and not on_com:
            status = "🔴 Flagged"
        elif connected_nodes == 0:
            status = "⚠️ Partial"
        elif desc_source == "none":
            status = "⚠️ Partial"
        elif jira_count == 0 and not has_desc:
            status = "⚠️ Partial"
        else:
            status = "✅ Good"

        rows.append({
            "name":           name,
            "services":       c["services"],
            "module":         module,
            "in_present":     in_present,
            "com_present":    com_present,
            "data_type":      data_type,
            "description":    desc,
            "desc_source":    desc_source,
            "jira_count":     jira_count,
            "config_page":    "✅" if wiki["config_page"] else "—",
            "module_page":    "✅" if wiki["module_page"] else "—",
            "cross_modules":  wiki["cross_module_pages"],
            "connected_nodes":connected_nodes,
            "status":         status,
        })

    # ── Summary stats ───────────────────────────────────────────────────────
    total     = len(rows)
    flagged   = sum(1 for r in rows if r["status"].startswith("🔴"))
    partial   = sum(1 for r in rows if r["status"].startswith("⚠️"))
    good      = sum(1 for r in rows if r["status"].startswith("✅"))
    with_jira = sum(1 for r in rows if r["jira_count"] > 0)

    print(f"\nSummary:")
    print(f"  Total: {total} | Good: {good} | Partial: {partial} | Flagged: {flagged}")
    print(f"  Configs with Jira mentions: {with_jira}")

    # ── Write Excel ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: All configs ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Config Validation"

    headers = [
        "Property Name", "Service(s)", "Module", ".in", ".com",
        "Data Type", "Description", "Desc Source",
        "Jira Tickets", "Config Page", "Module Page",
        "Cross-Module Pages", "Connected Nodes", "Status",
    ]

    # Header row
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = GRAY
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN

    # Data rows
    for row_idx, r in enumerate(rows, start=2):
        vals = [
            r["name"], r["services"], r["module"],
            r["in_present"], r["com_present"], r["data_type"],
            r["description"], r["desc_source"],
            r["jira_count"],
            r["config_page"], r["module_page"],
            r["cross_modules"], r["connected_nodes"],
            r["status"],
        ]
        if r["status"].startswith("🔴"):
            fill = RED
        elif r["status"].startswith("⚠️"):
            fill = YELLOW
        else:
            fill = GREEN

        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            # Only color the status and name columns, light for desc source
            if col_idx in (1, 14):  # name and status
                cell.fill = fill
            elif col_idx == 8 and r["desc_source"] == "none":  # desc source
                cell.fill = RED

    # Column widths
    col_widths = [40, 30, 22, 6, 6, 12, 60, 18, 10, 12, 12, 14, 14, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Sheet 2: Summary ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ["Metric", "Count", "Percent"],
        ["Total Configs", total, "100%"],
        ["✅ Good (full coverage)", good, f"{good/total*100:.1f}%"],
        ["⚠️ Partial (some coverage)", partial, f"{partial/total*100:.1f}%"],
        ["🔴 Flagged (no description or missing)", flagged, f"{flagged/total*100:.1f}%"],
        ["", "", ""],
        ["With Jira ticket mentions", with_jira, f"{with_jira/total*100:.1f}%"],
        ["No Jira mentions", total-with_jira, f"{(total-with_jira)/total*100:.1f}%"],
        ["", "", ""],
        ["On .in server", sum(1 for r in rows if r["in_present"]=="✅"), ""],
        ["On .com server", sum(1 for r in rows if r["com_present"]=="✅"), ""],
        ["On both servers", sum(1 for r in rows if r["in_present"]=="✅" and r["com_present"]=="✅"), ""],
        ["On .in only", sum(1 for r in rows if r["in_present"]=="✅" and r["com_present"]=="—"), ""],
        ["On .com only", sum(1 for r in rows if r["in_present"]=="—" and r["com_present"]=="✅"), ""],
        ["", "", ""],
        ["Desc from .com", sum(1 for r in rows if r["desc_source"]==".com"), ""],
        ["Desc from .in", sum(1 for r in rows if r["desc_source"]==".in"), ""],
        ["Desc from unique_configs", sum(1 for r in rows if r["desc_source"]=="unique_configs"), ""],
        ["Desc from pms_desc_cleaned", sum(1 for r in rows if r["desc_source"]=="pms_desc_cleaned"), ""],
        ["No description", sum(1 for r in rows if r["desc_source"]=="none"), ""],
    ]
    for r_idx, row_data in enumerate(summary_data, start=1):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN
            if r_idx == 1:
                cell.fill = GRAY
                cell.font = BOLD
            elif isinstance(val, str) and val.startswith("✅"):
                cell.fill = GREEN
            elif isinstance(val, str) and val.startswith("⚠️"):
                cell.fill = YELLOW
            elif isinstance(val, str) and val.startswith("🔴"):
                cell.fill = RED

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    # ── Sheet 3: Flagged only ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Flagged Configs")
    flagged_rows = [r for r in rows if r["status"].startswith("🔴")]
    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = GRAY
        cell.font = BOLD
        cell.border = THIN
    for row_idx, r in enumerate(flagged_rows, start=2):
        vals = [
            r["name"], r["services"], r["module"],
            r["in_present"], r["com_present"], r["data_type"],
            r["description"], r["desc_source"],
            r["jira_count"],
            r["config_page"], r["module_page"],
            r["cross_modules"], r["connected_nodes"],
            r["status"],
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN
            cell.fill = RED
    for i, w in enumerate(col_widths, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Sheet 4: Per-module breakdown ───────────────────────────────────────
    ws4 = wb.create_sheet("By Module")
    from collections import defaultdict
    module_stats = defaultdict(lambda: {"total": 0, "good": 0, "partial": 0, "flagged": 0})
    for r in rows:
        m = r["module"] or "unknown"
        module_stats[m]["total"] += 1
        if r["status"].startswith("✅"):
            module_stats[m]["good"] += 1
        elif r["status"].startswith("⚠️"):
            module_stats[m]["partial"] += 1
        else:
            module_stats[m]["flagged"] += 1

    mod_headers = ["Module", "Total", "✅ Good", "⚠️ Partial", "🔴 Flagged", "Coverage %"]
    for col, h in enumerate(mod_headers, start=1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = GRAY
        cell.font = BOLD
        cell.border = THIN

    for row_idx, (mod, stats) in enumerate(
        sorted(module_stats.items(), key=lambda x: -x[1]["flagged"]), start=2
    ):
        pct = f"{stats['good']/stats['total']*100:.0f}%" if stats["total"] else "—"
        vals = [mod, stats["total"], stats["good"], stats["partial"], stats["flagged"], pct]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN
        # color the row based on worst status
        row_fill = GREEN if stats["flagged"] == 0 and stats["partial"] == 0 else (
            RED if stats["flagged"] > 0 else YELLOW
        )
        ws4.cell(row=row_idx, column=1).fill = row_fill

    ws4.column_dimensions["A"].width = 30
    for i in range(2, 7):
        ws4.column_dimensions[get_column_letter(i)].width = 14

    # ── Save ────────────────────────────────────────────────────────────────
    wb.save(str(OUTPUT))
    print(f"\nReport written to: {OUTPUT}")
    print(f"  Sheets: Config Validation | Summary | Flagged Configs | By Module")


if __name__ == "__main__":
    build_report()
