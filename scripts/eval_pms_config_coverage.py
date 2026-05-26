#!/usr/bin/env python3
"""
Evaluate PMS config answer readiness across .in and .com workbooks.

This is a coverage/evidence evaluator, not a simple description checker. For
each config row it measures:
  - how much wiki/docs context exists beyond the generated config inventory
  - how many Jira tickets mention the config directly
  - whether related configs/dependency signals are discoverable
  - whether the config should be flagged for enrichment

Optional Claude mode can ask Claude to judge the collected evidence, but the
default run is deterministic and can complete the full inventory without model
cost.

Examples:
  python scripts/eval_pms_config_coverage.py

  python scripts/eval_pms_config_coverage.py --limit 20

  python scripts/eval_pms_config_coverage.py --ask-claude --limit 5
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sqlite3
import subprocess
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover - command-line guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install with `pip install -r requirements.txt`."
    ) from exc


ROOT = Path(__file__).resolve().parent.parent
WIKI = ROOT / "wiki"
DOCS = ROOT / "docs"
RAW_MODULES = ROOT / "raw" / "modules"
JIRA_DB = ROOT / "raw" / "jira" / "tickets.sqlite"

DEFAULT_IN_XLSX = Path("/Users/rudrakhare/Downloads/All WIS CONFIGS.xlsx")
DEFAULT_COM_XLSX = Path("/Users/rudrakhare/Downloads/wis_service_configs (1).xlsx")

GENERATED_CONFIG_PAGES = {
    "app-server-config.md",
    "booking-rule-engine.md",
    "emp-experience-common.md",
    "emp-experience-email.md",
    "emp-experience-internal.md",
    "guard-app.md",
    "meeting-rooms.md",
    "mobile-app-server.md",
    "pms.md",
    "visitor-management.md",
    "wis-seat-booking.md",
}

TEXT_EXTS = {
    ".md",
    ".txt",
    ".csv",
    ".tsv",
    ".json",
    ".yaml",
    ".yml",
    ".toml",
}

STOPWORDS = {
    "able",
    "about",
    "above",
    "across",
    "action",
    "active",
    "admin",
    "after",
    "allow",
    "allows",
    "also",
    "and",
    "app",
    "are",
    "based",
    "before",
    "booking",
    "bookings",
    "button",
    "can",
    "check",
    "config",
    "configuration",
    "configured",
    "controls",
    "created",
    "defines",
    "disable",
    "disabled",
    "does",
    "employee",
    "enable",
    "enabled",
    "enables",
    "false",
    "feature",
    "flow",
    "for",
    "from",
    "has",
    "into",
    "kiosk",
    "list",
    "management",
    "minutes",
    "module",
    "not",
    "notification",
    "office",
    "only",
    "option",
    "page",
    "property",
    "request",
    "resource",
    "same",
    "send",
    "sent",
    "service",
    "set",
    "should",
    "show",
    "specific",
    "the",
    "this",
    "time",
    "true",
    "type",
    "used",
    "user",
    "value",
    "via",
    "when",
    "where",
    "which",
    "with",
    "workflow",
}

BEHAVIOR_WORDS = {
    "allow",
    "allows",
    "booking",
    "block",
    "cancel",
    "checkin",
    "check-in",
    "controls",
    "disable",
    "disabled",
    "enable",
    "enabled",
    "enables",
    "hide",
    "notification",
    "release",
    "restrict",
    "scan",
    "send",
    "show",
    "visible",
}

SCOPE_WORDS = {
    ".com",
    ".in",
    "buid",
    "client",
    "employee",
    "global",
    "office",
    "org",
    "role",
    "server",
    "tenant",
    "user",
}

DEPENDENCY_RE = re.compile(
    r"(?i)\b("
    r"requires?|depends?|dependent|dependency|coupled|must be enabled|"
    r"along with|only applies|has no effect|master property|pre[- ]?requisite|"
    r"enable[d]? with|if .{0,80} then"
    r")\b"
)

NOISE_PRONE_NAMES = {
    "analytics",
    "radius",
    "districts",
    "sequence",
    "timezone",
    "location",
    "language",
    "role",
    "status",
}


@dataclass
class ConfigRecord:
    record_id: str
    server: str
    workbook_path: Path
    sheet: str
    row_number: int
    property_name: str
    description: str = ""
    data_type: str = ""
    output: dict[str, Any] = field(default_factory=dict)


@dataclass
class TextMatch:
    path: Path
    line_no: int
    line: str


@dataclass
class CorpusFile:
    path: Path
    text: str
    lower: str


@dataclass
class Ticket:
    idx: int
    key: str
    project: str
    status: str
    status_category: str
    priority: str
    functional_area: str
    updated_at: str
    summary: str
    description_text: str
    comments_text: str
    combined: str
    combined_lower: str
    tokens: set[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--in-xlsx", type=Path, default=DEFAULT_IN_XLSX)
    parser.add_argument("--com-xlsx", type=Path, default=DEFAULT_COM_XLSX)
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=ROOT / "eval_runs" / f"pms_config_eval_{timestamp_slug()}",
    )
    parser.add_argument("--jira-db", type=Path, default=JIRA_DB)
    parser.add_argument("--limit", type=int, help="Evaluate only the first N config rows.")
    parser.add_argument(
        "--server",
        choices=["in", "com", "both"],
        default="both",
        help="Restrict evaluation to one server workbook.",
    )
    parser.add_argument(
        "--ask-claude",
        action="store_true",
        help="Ask Claude to judge the collected evidence for each evaluated row.",
    )
    parser.add_argument("--claude-bin", default="claude")
    parser.add_argument("--claude-model", default="sonnet")
    parser.add_argument("--claude-timeout-s", type=int, default=180)
    parser.add_argument("--sleep-ms", type=int, default=0)
    parser.add_argument(
        "--max-ticket-snippets",
        type=int,
        default=5,
        help="Top exact Jira tickets to expose in output snippets.",
    )
    return parser.parse_args()


def timestamp_slug() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_meaningful_description(text: str) -> bool:
    s = clean(text)
    if not s or s in {"-", "—", "N/A", "NA", "na", "n/a"}:
        return False
    return len(re.sub(r"\W+", "", s)) >= 12


def split_aliases(property_name: str) -> list[str]:
    raw_parts = re.split(r"\s*/\s*|\s*,\s*|\s+\|\s+|\s+or\s+", property_name)
    aliases = []
    for part in raw_parts:
        part = part.strip(" `\"'")
        if part:
            aliases.append(part)
    if property_name not in aliases:
        aliases.insert(0, property_name)
    return list(dict.fromkeys(aliases))


def camel_to_words(text: str) -> str:
    text = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", text)
    text = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", text)
    return text


def tokenize(text: str) -> list[str]:
    text = camel_to_words(text)
    tokens = re.findall(r"[A-Za-z0-9_]+", text.lower())
    return [t for t in tokens if len(t) >= 3 and t not in STOPWORDS]


def candidate_tokens(text: str) -> list[str]:
    """Tokens for narrowing exact search candidates.

    This intentionally keeps stopwords because some config names are generic
    role labels like "GLOBAL ADMIN"; stopwords are bad for semantic search but
    useful for reducing the exact-match ticket set.
    """
    text = camel_to_words(text)
    return list(dict.fromkeys(t for t in re.findall(r"[A-Za-z0-9_]+", text.lower()) if len(t) >= 3))


def query_tokens(record: ConfigRecord) -> list[str]:
    parts = [record.property_name, record.description, record.sheet]
    tokens = tokenize(" ".join(parts))
    # Keep order but remove duplicates.
    return list(dict.fromkeys(tokens))[:24]


def find_header(ws: Any) -> tuple[int, dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12)):
        values = [clean(cell.value) for cell in row]
        lowered = [v.lower() for v in values]
        if "property name" in lowered:
            header_row = row[0].row
            return header_row, {value: idx + 1 for idx, value in enumerate(values) if value}
    raise ValueError(f"Could not find 'Property Name' header in sheet {ws.title!r}")


def load_inventory(path: Path, server: str) -> tuple[list[ConfigRecord], dict[str, list[ConfigRecord]]]:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[ConfigRecord] = []
    by_property: dict[str, list[ConfigRecord]] = defaultdict(list)
    try:
        for ws in workbook.worksheets:
            header_row, cols = find_header(ws)
            property_col = cols.get("Property Name")
            description_col = cols.get("Description")
            data_type_col = cols.get("Data Type")
            if not property_col:
                continue
            for row_number in range(header_row + 1, ws.max_row + 1):
                property_name = clean(ws.cell(row=row_number, column=property_col).value)
                if not property_name:
                    continue
                data_type = clean(ws.cell(row=row_number, column=data_type_col).value) if data_type_col else ""
                description = (
                    clean(ws.cell(row=row_number, column=description_col).value)
                    if description_col
                    else ""
                )
                rec = ConfigRecord(
                    record_id=f"{server}:{ws.title}:{row_number}:{property_name}",
                    server=server,
                    workbook_path=path,
                    sheet=ws.title,
                    row_number=row_number,
                    property_name=property_name,
                    description=description,
                    data_type=data_type,
                )
                records.append(rec)
                by_property[property_name.lower()].append(rec)
    finally:
        workbook.close()
    return records, by_property


def iter_corpus_files() -> list[CorpusFile]:
    files: list[CorpusFile] = []
    for scan_dir in (WIKI, DOCS, RAW_MODULES, ROOT / "config"):
        if not scan_dir.exists():
            continue
        for path in scan_dir.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in TEXT_EXTS:
                continue
            if "/raw/jira/" in str(path):
                continue
            try:
                text = path.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
            files.append(CorpusFile(path=path, text=text, lower=text.lower()))
    return files


def is_wiki_path(path: Path) -> bool:
    return WIKI in path.parents


def is_rich_context_path(path: Path) -> bool:
    rel = short_path(path)
    if rel.startswith("wiki/sources/pms-configs-"):
        return False
    if rel.startswith("raw/modules/pms-configs-"):
        return False
    if rel.startswith("wiki/configs/") and path.name in GENERATED_CONFIG_PAGES:
        return False
    return True


def exact_match_pattern(alias: str) -> re.Pattern[str]:
    escaped = re.escape(alias)
    if re.fullmatch(r"[A-Za-z0-9_]+", alias):
        return re.compile(rf"(?<![A-Za-z0-9_]){escaped}(?![A-Za-z0-9_])", re.IGNORECASE)
    return re.compile(escaped, re.IGNORECASE)


def find_text_matches(record: ConfigRecord, files: list[CorpusFile]) -> list[TextMatch]:
    patterns = [exact_match_pattern(alias) for alias in split_aliases(record.property_name)]
    matches: list[TextMatch] = []
    for file in files:
        if not any(pattern.search(file.text) for pattern in patterns):
            continue
        for line_no, line in enumerate(file.text.splitlines(), start=1):
            if any(pattern.search(line) for pattern in patterns):
                matches.append(TextMatch(path=file.path, line_no=line_no, line=line.strip()))
    return matches


def load_tickets(db_path: Path) -> list[Ticket]:
    if not db_path.exists():
        return []
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT key, project, status, status_category, priority, functional_area,
                   updated_at, summary, description_text, comments_text
            FROM tickets
            """
        ).fetchall()
    finally:
        conn.close()

    tickets: list[Ticket] = []
    for idx, row in enumerate(rows):
        combined = "\n".join(
            clean(row[col])
            for col in ("summary", "description_text", "comments_text")
            if clean(row[col])
        )
        combined_for_tokens = combined[:60000]
        tickets.append(
            Ticket(
                idx=idx,
                key=clean(row["key"]),
                project=clean(row["project"]),
                status=clean(row["status"]),
                status_category=clean(row["status_category"]),
                priority=clean(row["priority"]),
                functional_area=clean(row["functional_area"]),
                updated_at=clean(row["updated_at"]),
                summary=clean(row["summary"]),
                description_text=clean(row["description_text"]),
                comments_text=clean(row["comments_text"]),
                combined=combined,
                combined_lower=combined.lower(),
                tokens=set(tokenize(combined_for_tokens)),
            )
        )
    return tickets


def build_token_index(tickets: list[Ticket]) -> dict[str, set[int]]:
    index: dict[str, set[int]] = defaultdict(set)
    for ticket in tickets:
        for token in ticket.tokens:
            index[token].add(ticket.idx)
    return index


def jira_exact_matches(
    record: ConfigRecord,
    tickets: list[Ticket],
    token_index: dict[str, set[int]],
) -> list[Ticket]:
    patterns = [exact_match_pattern(alias) for alias in split_aliases(record.property_name)]
    candidate_indexes: set[int] = set()
    for alias in split_aliases(record.property_name):
        alias_tokens = candidate_tokens(alias)
        token_sets = [token_index.get(token, set()) for token in alias_tokens if token in token_index]
        if not token_sets:
            continue
        # Intersect terms from the same alias to reduce noisy scans. For a
        # single-token alias this is just that token's posting list.
        alias_candidates = set.intersection(*token_sets)
        candidate_indexes.update(alias_candidates)
    if not candidate_indexes:
        candidate_indexes = set(range(len(tickets)))

    matched: list[Ticket] = []
    for idx in candidate_indexes:
        ticket = tickets[idx]
        if any(pattern.search(ticket.combined) for pattern in patterns):
            matched.append(ticket)
    return sorted(
        matched,
        key=lambda t: (t.status_category != "done", t.updated_at),
        reverse=True,
    )


def jira_keyword_matches(
    record: ConfigRecord,
    tickets: list[Ticket],
    token_index: dict[str, set[int]],
    exact_ticket_keys: set[str],
) -> list[tuple[Ticket, int]]:
    tokens = query_tokens(record)
    if not tokens:
        return []

    # Prefer more selective terms so generic words do not swamp the search.
    token_dfs = [(token, len(token_index.get(token, set()))) for token in tokens]
    selective = [t for t, df in sorted(token_dfs, key=lambda item: item[1]) if 0 < df <= 4500]
    selected_tokens = selective[:10]
    if not selected_tokens:
        return []

    candidates: set[int] = set()
    for token in selected_tokens:
        candidates.update(token_index.get(token, set()))

    scored: list[tuple[Ticket, int]] = []
    service_tokens = set(tokenize(record.sheet))
    for idx in candidates:
        ticket = tickets[idx]
        if ticket.key in exact_ticket_keys:
            continue
        token_hits = sum(1 for token in selected_tokens if token in ticket.tokens)
        service_hits = sum(1 for token in service_tokens if token in ticket.tokens)
        score = token_hits + min(service_hits, 2)
        if record.data_type and record.data_type.lower() in ticket.combined_lower:
            score += 1
        if score >= 4:
            scored.append((ticket, score))

    return sorted(scored, key=lambda item: (item[1], item[0].updated_at), reverse=True)[:25]


def ticket_snippet(ticket: Ticket, aliases: list[str], max_len: int = 260) -> str:
    haystacks = [
        ("summary", ticket.summary),
        ("description", ticket.description_text),
        ("comments", ticket.comments_text),
    ]
    lowered_aliases = [a.lower() for a in aliases if a]
    for label, text in haystacks:
        lower = text.lower()
        positions = [lower.find(alias) for alias in lowered_aliases if lower.find(alias) >= 0]
        if positions:
            pos = min(positions)
            start = max(0, pos - 90)
            end = min(len(text), pos + max_len)
            snippet = compact(text[start:end])
            return f"{ticket.key} {label}: {snippet}"
    return f"{ticket.key} summary: {compact(ticket.summary)[:max_len]}"


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", clean(text))


def short_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def path_ref(match: TextMatch) -> str:
    return f"{short_path(match.path)}:{match.line_no}"


def sample_wiki_refs(matches: list[TextMatch], limit: int = 5) -> str:
    ordered = sorted(
        matches,
        key=lambda m: (not is_rich_context_path(m.path), short_path(m.path), m.line_no),
    )
    refs = []
    seen = set()
    for match in ordered:
        ref = path_ref(match)
        if ref in seen:
            continue
        refs.append(ref)
        seen.add(ref)
        if len(refs) >= limit:
            break
    return ", ".join(refs)


def sample_lines(matches: list[TextMatch], limit: int = 5) -> list[str]:
    ordered = sorted(
        matches,
        key=lambda m: (not is_rich_context_path(m.path), short_path(m.path), m.line_no),
    )
    snippets = []
    seen = set()
    for match in ordered:
        ref = path_ref(match)
        if ref in seen:
            continue
        snippets.append(f"{ref}: {compact(match.line)[:320]}")
        seen.add(ref)
        if len(snippets) >= limit:
            break
    return snippets


def find_connected_configs(
    evidence_text: str,
    all_config_names: list[str],
    current_aliases: list[str],
    limit: int = 18,
) -> list[str]:
    lower = evidence_text.lower()
    current = {alias.lower() for alias in current_aliases}
    found = []
    for name in all_config_names:
        aliases = split_aliases(name)
        if any(alias.lower() in current for alias in aliases):
            continue
        if all(len(alias) < 5 for alias in aliases):
            continue
        for alias in aliases:
            if len(alias) < 5:
                continue
            if exact_match_pattern(alias).search(evidence_text):
                found.append(name)
                break
        if len(found) >= limit:
            break
    return list(dict.fromkeys(found))


def find_dependency_signals(snippets: list[str], limit: int = 4) -> list[str]:
    signals = []
    for snippet in snippets:
        if DEPENDENCY_RE.search(snippet):
            signals.append(snippet[:360])
        if len(signals) >= limit:
            break
    return signals


def has_any_word(text: str, words: set[str]) -> bool:
    lower = text.lower()
    return any(word in lower for word in words)


def evaluate_record(
    record: ConfigRecord,
    *,
    other_server_lookup: dict[str, list[ConfigRecord]],
    all_config_names: list[str],
    corpus_files: list[CorpusFile],
    tickets: list[Ticket],
    token_index: dict[str, set[int]],
    max_ticket_snippets: int,
    ask_claude: bool,
    claude_bin: str,
    claude_model: str,
    claude_timeout_s: int,
) -> dict[str, Any]:
    aliases = split_aliases(record.property_name)
    text_matches = find_text_matches(record, corpus_files)
    wiki_matches = [m for m in text_matches if is_wiki_path(m.path)]
    rich_matches = [m for m in wiki_matches if is_rich_context_path(m.path)]
    raw_matches = [m for m in text_matches if "/raw/" in str(m.path)]

    exact_tickets = jira_exact_matches(record, tickets, token_index)
    exact_keys = {ticket.key for ticket in exact_tickets}
    keyword_tickets = jira_keyword_matches(record, tickets, token_index, exact_keys)

    wiki_snippets = sample_lines(wiki_matches, limit=6)
    ticket_snippets = [
        ticket_snippet(ticket, aliases) for ticket in exact_tickets[:max_ticket_snippets]
    ]
    keyword_ticket_snippets = [
        f"{ticket.key} score={score}: {compact(ticket.summary)[:240]}"
        for ticket, score in keyword_tickets[:5]
    ]
    evidence_text = "\n".join(
        [
            record.description,
            "\n".join(wiki_snippets),
            "\n".join(ticket_snippets),
            "\n".join(keyword_ticket_snippets),
        ]
    )

    connected_configs = find_connected_configs(evidence_text, all_config_names, aliases)
    dependency_signals = find_dependency_signals(wiki_snippets + ticket_snippets)
    present_other = other_server_lookup.get(record.property_name.lower(), [])
    present_other_sheets = sorted({r.sheet for r in present_other})

    meaningful_desc = is_meaningful_description(record.description)
    has_type = bool(record.data_type and record.data_type not in {"-", "—", "N/A"})
    has_behavior = has_any_word(evidence_text, BEHAVIOR_WORDS)
    has_scope = has_any_word(evidence_text, SCOPE_WORDS)
    has_default = bool(re.search(r"(?i)\b(default|true|false|null|empty|blank)\b", evidence_text))
    has_dependency = bool(dependency_signals or connected_configs)
    has_operational = len(exact_tickets) > 0
    has_rich_wiki = len(rich_matches) > 0

    dimensions = {
        "purpose": meaningful_desc or has_rich_wiki or has_operational,
        "data_type": has_type or record.server == "in",  # .in source has no Data Type column.
        "behavior": has_behavior,
        "scope": has_scope,
        "default_or_values": has_default,
        "dependencies": has_dependency,
        "jira_examples": has_operational,
        "rich_wiki_context": has_rich_wiki,
    }
    weights = {
        "purpose": 15,
        "data_type": 10,
        "behavior": 15,
        "scope": 10,
        "default_or_values": 10,
        "dependencies": 15,
        "jira_examples": 15,
        "rich_wiki_context": 10,
    }
    coverage_score = sum(weights[name] for name, ok in dimensions.items() if ok)
    missing_dimensions = [name for name, ok in dimensions.items() if not ok]

    noise_risk = (
        record.property_name.lower() in NOISE_PRONE_NAMES
        or (len(record.property_name) < 8 and len(exact_tickets) > 30)
        or len(exact_tickets) > 120
    )

    confidence = "HIGH"
    if coverage_score < 30:
        confidence = "UNKNOWN"
    elif coverage_score < 55:
        confidence = "LOW"
    elif coverage_score < 75:
        confidence = "MEDIUM"
    if noise_risk:
        confidence = "REVIEW"

    flag_reasons = []
    if not meaningful_desc and not has_operational and not has_rich_wiki:
        flag_reasons.append("No meaningful source description, rich wiki context, or exact Jira evidence.")
    if not has_operational and not has_rich_wiki:
        flag_reasons.append("Only inventory/raw context found; no operational Jira or rich wiki evidence.")
    if meaningful_desc and not has_operational and not has_rich_wiki:
        flag_reasons.append("Only a one-line description is available; user enablement/dependency questions are not evidence-backed.")
    if confidence in {"UNKNOWN", "LOW", "REVIEW"}:
        flag_reasons.append(f"Coverage confidence is {confidence}.")
    if noise_risk:
        flag_reasons.append("Exact Jira matches look noisy or generic; needs manual curation.")

    flagged = bool(flag_reasons)
    if not flagged and "dependencies" in missing_dimensions and record.data_type.upper() == "BOOLEAN":
        # Do not hard-fail otherwise good configs, but surface the gap.
        flag_reasons.append("Dependency/connected-config context is not explicit.")

    recommended_action = recommended_next_action(
        flagged=flagged,
        missing_dimensions=missing_dimensions,
        has_operational=has_operational,
        has_rich_wiki=has_rich_wiki,
        meaningful_desc=meaningful_desc,
    )

    row = {
        "eval_flagged": "YES" if flagged else "NO",
        "eval_confidence": confidence,
        "eval_coverage_score": coverage_score,
        "eval_flag_reason": " ".join(flag_reasons[:3]) if flag_reasons else "",
        "eval_missing_dimensions": ", ".join(missing_dimensions),
        "eval_recommended_action": recommended_action,
        "eval_wiki_match_count": len(wiki_matches),
        "eval_wiki_rich_match_count": len(rich_matches),
        "eval_raw_match_count": len(raw_matches),
        "eval_wiki_refs": sample_wiki_refs(wiki_matches),
        "eval_jira_exact_ticket_count": len(exact_tickets),
        "eval_jira_keyword_ticket_count": len(keyword_tickets),
        "eval_jira_top_keys": ", ".join(ticket.key for ticket in exact_tickets[:8]),
        "eval_jira_keyword_keys": ", ".join(ticket.key for ticket, _ in keyword_tickets[:8]),
        "eval_jira_snippets": " | ".join(ticket_snippets[:3]),
        "eval_connected_config_count": len(connected_configs),
        "eval_connected_configs": ", ".join(connected_configs),
        "eval_dependency_signals": " | ".join(dependency_signals),
        "eval_present_on_other_server": "YES" if present_other else "NO",
        "eval_other_server_sheets": ", ".join(present_other_sheets),
    }

    if ask_claude:
        claude_result = ask_claude_to_judge(
            record=record,
            deterministic_row=row,
            wiki_snippets=wiki_snippets,
            ticket_snippets=ticket_snippets,
            keyword_ticket_snippets=keyword_ticket_snippets,
            claude_bin=claude_bin,
            claude_model=claude_model,
            timeout_s=claude_timeout_s,
        )
        row.update(claude_result)

    return row


def recommended_next_action(
    *,
    flagged: bool,
    missing_dimensions: list[str],
    has_operational: bool,
    has_rich_wiki: bool,
    meaningful_desc: bool,
) -> str:
    if not meaningful_desc and not has_operational and not has_rich_wiki:
        return "Ask owning team for purpose/type/scope; then create a config intelligence entry."
    if has_operational and not has_rich_wiki:
        return "Promote Jira evidence into wiki: behavior, scope, examples, and known edge cases."
    if has_rich_wiki and not has_operational:
        return "Search/curate Jira examples or explicitly document that no ticket evidence was found."
    if "dependencies" in missing_dimensions:
        return "Document dependent configs, prerequisites, and safe enablement sequence."
    if "default_or_values" in missing_dimensions:
        return "Document default value, valid values, and rollback behavior."
    if flagged:
        return "Review missing dimensions and add evidence-backed wiki notes."
    return "No immediate action; keep evidence fresh as Jira/docs change."


def claude_json_schema() -> str:
    schema = {
        "type": "object",
        "properties": {
            "answer_readiness": {"type": "string", "enum": ["ready", "partial", "not_ready"]},
            "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
            "summary": {"type": "string"},
            "can_answer": {"type": "array", "items": {"type": "string"}},
            "missing_info": {"type": "array", "items": {"type": "string"}},
            "should_flag": {"type": "boolean"},
            "flag_reason": {"type": "string"},
        },
        "required": [
            "answer_readiness",
            "confidence",
            "summary",
            "can_answer",
            "missing_info",
            "should_flag",
            "flag_reason",
        ],
        "additionalProperties": False,
    }
    return json.dumps(schema)


def ask_claude_to_judge(
    *,
    record: ConfigRecord,
    deterministic_row: dict[str, Any],
    wiki_snippets: list[str],
    ticket_snippets: list[str],
    keyword_ticket_snippets: list[str],
    claude_bin: str,
    claude_model: str,
    timeout_s: int,
) -> dict[str, Any]:
    prompt = f"""
You are evaluating whether internal teams can answer user questions about a PMS config
from the available evidence. Do not invent missing facts. Judge only the evidence below.

Config:
- server: {record.server}
- service/sheet: {record.sheet}
- property: {record.property_name}
- data_type: {record.data_type or "not in source"}
- source_description: {record.description or "not documented"}

Deterministic evidence metrics:
{json.dumps(deterministic_row, indent=2)}

Wiki snippets:
{json.dumps(wiki_snippets, indent=2)}

Exact Jira snippets:
{json.dumps(ticket_snippets, indent=2)}

Keyword Jira snippets:
{json.dumps(keyword_ticket_snippets, indent=2)}

Evaluate whether a support/product/user-facing answer about this config would be good
enough for:
1. what the config does,
2. how/when to enable it,
3. dependencies or connected configs,
4. scope/default/type,
5. Jira/customer context.
Return JSON only.
""".strip()
    cmd = [
        claude_bin,
        "-p",
        "--model",
        claude_model,
        "--output-format",
        "json",
        "--json-schema",
        claude_json_schema(),
        "--tools",
        "",
        "--permission-mode",
        "dontAsk",
        prompt,
    ]
    started = time.time()
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(ROOT),
            text=True,
            capture_output=True,
            timeout=timeout_s,
            check=False,
        )
    except Exception as exc:  # noqa: BLE001
        return {
            "claude_eval_status": "ERROR",
            "claude_error": str(exc),
        }

    elapsed_ms = int((time.time() - started) * 1000)
    if proc.returncode != 0:
        return {
            "claude_eval_status": "ERROR",
            "claude_error": compact(proc.stderr or proc.stdout)[:1200],
            "claude_elapsed_ms": elapsed_ms,
        }

    try:
        outer = json.loads(proc.stdout)
        payload = outer.get("result", outer)
        if isinstance(payload, str):
            payload = json.loads(payload)
    except Exception as exc:  # noqa: BLE001
        return {
            "claude_eval_status": "PARSE_ERROR",
            "claude_error": f"{exc}: {compact(proc.stdout)[:1200]}",
            "claude_elapsed_ms": elapsed_ms,
        }

    return {
        "claude_eval_status": "OK",
        "claude_answer_readiness": payload.get("answer_readiness", ""),
        "claude_confidence": payload.get("confidence", ""),
        "claude_summary": payload.get("summary", ""),
        "claude_can_answer": " | ".join(payload.get("can_answer", []) or []),
        "claude_missing_info": " | ".join(payload.get("missing_info", []) or []),
        "claude_should_flag": "YES" if payload.get("should_flag") else "NO",
        "claude_flag_reason": payload.get("flag_reason", ""),
        "claude_elapsed_ms": elapsed_ms,
    }


def output_columns(include_claude: bool) -> list[str]:
    cols = [
        "eval_flagged",
        "eval_confidence",
        "eval_coverage_score",
        "eval_flag_reason",
        "eval_missing_dimensions",
        "eval_recommended_action",
        "eval_wiki_match_count",
        "eval_wiki_rich_match_count",
        "eval_raw_match_count",
        "eval_wiki_refs",
        "eval_jira_exact_ticket_count",
        "eval_jira_keyword_ticket_count",
        "eval_jira_top_keys",
        "eval_jira_keyword_keys",
        "eval_jira_snippets",
        "eval_connected_config_count",
        "eval_connected_configs",
        "eval_dependency_signals",
        "eval_present_on_other_server",
        "eval_other_server_sheets",
    ]
    if include_claude:
        cols.extend(
            [
                "claude_eval_status",
                "claude_answer_readiness",
                "claude_confidence",
                "claude_summary",
                "claude_can_answer",
                "claude_missing_info",
                "claude_should_flag",
                "claude_flag_reason",
                "claude_elapsed_ms",
                "claude_error",
            ]
        )
    return cols


def write_evaluated_workbook(
    input_path: Path,
    output_path: Path,
    records: list[ConfigRecord],
    include_claude: bool,
) -> None:
    wb = load_workbook(input_path)
    by_sheet_row = {(r.sheet, r.row_number): r for r in records}
    columns = output_columns(include_claude)
    fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        header_row, _ = find_header(ws)
        start_col = ws.max_column + 1
        for offset, col_name in enumerate(columns):
            cell = ws.cell(row=header_row, column=start_col + offset, value=col_name)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row_number in range(header_row + 1, ws.max_row + 1):
            record = by_sheet_row.get((ws.title, row_number))
            if not record:
                continue
            for offset, col_name in enumerate(columns):
                value = record.output.get(col_name, "")
                cell = ws.cell(row=row_number, column=start_col + offset, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def write_jsonl(path: Path, records: list[ConfigRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            payload = {
                "record_id": record.record_id,
                "server": record.server,
                "sheet": record.sheet,
                "row_number": record.row_number,
                "property_name": record.property_name,
                "data_type": record.data_type,
                "description": record.description,
                **record.output,
            }
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def write_csv(path: Path, records: list[ConfigRecord], include_claude: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "server",
        "sheet",
        "row_number",
        "property_name",
        "data_type",
        "description",
        *output_columns(include_claude),
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "server": record.server,
                    "sheet": record.sheet,
                    "row_number": record.row_number,
                    "property_name": record.property_name,
                    "data_type": record.data_type,
                    "description": record.description,
                    **record.output,
                }
            )


def write_summary(path: Path, records: list[ConfigRecord], outputs: dict[str, Path]) -> None:
    total = len(records)
    flag_counts = Counter(r.output.get("eval_flagged", "") for r in records)
    confidence_counts = Counter(r.output.get("eval_confidence", "") for r in records)
    server_counts = Counter(r.server for r in records)
    sheet_flag_counts: Counter[str] = Counter()
    reason_counts: Counter[str] = Counter()
    for r in records:
        if r.output.get("eval_flagged") == "YES":
            sheet_flag_counts[f"{r.server}:{r.sheet}"] += 1
            reason = r.output.get("eval_flag_reason", "").split(".")[0].strip()
            if reason:
                reason_counts[reason] += 1

    lines = [
        "# PMS Config Evaluation Summary",
        "",
        f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}",
        f"Total evaluated rows: **{total}**",
        "",
        "## Outputs",
        "",
    ]
    for label, output_path in outputs.items():
        lines.append(f"- {label}: `{short_path(output_path)}`")

    lines.extend(
        [
            "",
            "## Server Counts",
            "",
            "| Server | Rows |",
            "|---|---:|",
        ]
    )
    for server, count in sorted(server_counts.items()):
        lines.append(f"| `{server}` | {count} |")

    lines.extend(
        [
            "",
            "## Flag Counts",
            "",
            "| Flagged | Rows |",
            "|---|---:|",
        ]
    )
    for flag, count in sorted(flag_counts.items()):
        lines.append(f"| `{flag or 'blank'}` | {count} |")

    lines.extend(
        [
            "",
            "## Confidence Counts",
            "",
            "| Confidence | Rows |",
            "|---|---:|",
        ]
    )
    for confidence, count in sorted(confidence_counts.items()):
        lines.append(f"| `{confidence or 'blank'}` | {count} |")

    lines.extend(
        [
            "",
            "## Most Flagged Sheets",
            "",
            "| Server:Sheet | Flagged Rows |",
            "|---|---:|",
        ]
    )
    for sheet, count in sheet_flag_counts.most_common(20):
        lines.append(f"| `{sheet}` | {count} |")

    lines.extend(
        [
            "",
            "## Top Flag Reasons",
            "",
            "| Reason | Rows |",
            "|---|---:|",
        ]
    )
    for reason, count in reason_counts.most_common(20):
        lines.append(f"| {reason} | {count} |")

    lines.extend(
        [
            "",
            "## Rubric",
            "",
            "A row is flagged when it has weak answer readiness: missing meaningful description, "
            "no rich wiki context, no exact Jira evidence, low confidence, or noisy/generic Jira matches.",
            "",
            "Coverage dimensions are purpose, data type, behavior, scope, default/values, "
            "dependencies, Jira examples, and rich wiki context.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    print("Loading inventories...", flush=True)
    in_records, in_lookup = load_inventory(args.in_xlsx, "in")
    com_records, com_lookup = load_inventory(args.com_xlsx, "com")

    all_records: list[ConfigRecord] = []
    if args.server in {"in", "both"}:
        all_records.extend(in_records)
    if args.server in {"com", "both"}:
        all_records.extend(com_records)
    if args.limit:
        all_records = all_records[: args.limit]

    all_config_names = sorted(
        {r.property_name for r in [*in_records, *com_records]},
        key=lambda name: (-len(name), name.lower()),
    )

    print(f"Loading text corpus from wiki/docs/raw/config...", flush=True)
    corpus_files = iter_corpus_files()
    print(f"Loaded {len(corpus_files)} text files.", flush=True)

    print("Loading Jira tickets...", flush=True)
    tickets = load_tickets(args.jira_db)
    token_index = build_token_index(tickets)
    print(f"Loaded {len(tickets)} Jira tickets.", flush=True)

    print(f"Evaluating {len(all_records)} config rows...", flush=True)
    for idx, record in enumerate(all_records, start=1):
        other_lookup = com_lookup if record.server == "in" else in_lookup
        record.output = evaluate_record(
            record,
            other_server_lookup=other_lookup,
            all_config_names=all_config_names,
            corpus_files=corpus_files,
            tickets=tickets,
            token_index=token_index,
            max_ticket_snippets=args.max_ticket_snippets,
            ask_claude=args.ask_claude,
            claude_bin=args.claude_bin,
            claude_model=args.claude_model,
            claude_timeout_s=args.claude_timeout_s,
        )
        if idx % 25 == 0 or idx == len(all_records):
            print(
                f"  [{idx}/{len(all_records)}] {record.server}:{record.sheet}:{record.property_name}",
                flush=True,
            )
        if args.sleep_ms > 0:
            time.sleep(args.sleep_ms / 1000.0)

    # Only write evaluated workbook rows for the records included in this run.
    in_output = args.out_dir / f"{args.in_xlsx.stem}.evaluated.xlsx"
    com_output = args.out_dir / f"{args.com_xlsx.stem}.evaluated.xlsx"
    if args.server in {"in", "both"}:
        write_evaluated_workbook(
            args.in_xlsx,
            in_output,
            [r for r in all_records if r.server == "in"],
            args.ask_claude,
        )
    if args.server in {"com", "both"}:
        write_evaluated_workbook(
            args.com_xlsx,
            com_output,
            [r for r in all_records if r.server == "com"],
            args.ask_claude,
        )

    jsonl_output = args.out_dir / "pms_config_eval_rows.jsonl"
    csv_output = args.out_dir / "pms_config_eval_rows.csv"
    summary_output = args.out_dir / "summary.md"
    write_jsonl(jsonl_output, all_records)
    write_csv(csv_output, all_records, args.ask_claude)
    outputs = {
        "row jsonl": jsonl_output,
        "row csv": csv_output,
        "summary": summary_output,
    }
    if args.server in {"in", "both"}:
        outputs[".in workbook"] = in_output
    if args.server in {"com", "both"}:
        outputs[".com workbook"] = com_output
    write_summary(summary_output, all_records, outputs)

    print("\nDone.", flush=True)
    for label, path in outputs.items():
        print(f"- {label}: {path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
