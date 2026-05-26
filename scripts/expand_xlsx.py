#!/usr/bin/env python3
"""
expand_xlsx.py — Explode every .xlsx under raw/ into per-sheet CSV files.

For each <name>.xlsx found under raw/, creates a sibling directory <name>/
containing one <SheetName>.csv per non-empty sheet. Completely empty sheets
(no rows with data beyond the header) are skipped with a warning.

The original .xlsx is left in place — it is the source of truth.

Idempotent: CSVs that already exist with identical content (SHA-256) are not
rewritten. Safe to re-run after every sync.

Sheet name → filename mapping
------------------------------
  Sheet name "Connected Properties" → "Connected Properties.csv"
  Characters illegal on most filesystems ( / \\ : * ? " < > | ) are replaced
  with underscores. Trailing dots/spaces are stripped.

Examples
--------
  # Expand all xlsx under raw/ (default)
  python scripts/expand_xlsx.py

  # Target a specific directory
  python scripts/expand_xlsx.py --path raw/modules/_root

  # Preview without writing
  python scripts/expand_xlsx.py --dry-run

  # Also expand xlsx inside nested subdirectories of a module
  python scripts/expand_xlsx.py --recursive
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
from io import StringIO
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "ERROR: openpyxl is not installed.\n"
        "  Install with:  python3.11 -m pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SEARCH_ROOT = ROOT / "raw"

# Characters that are illegal in filenames on Windows / macOS / Linux
_ILLEGAL_CHARS = re.compile(r'[/\\:*?"<>|]')


def sanitize_sheet_name(name: str) -> str:
    """Convert a sheet name to a safe filename stem."""
    safe = _ILLEGAL_CHARS.sub("_", name).strip(". ")
    return safe or "sheet"


def sheet_to_csv_bytes(ws) -> bytes:
    """Render an openpyxl worksheet to UTF-8 CSV bytes."""
    buf = StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v) for v in row])
    return buf.getvalue().encode("utf-8")


def is_sheet_empty(ws) -> bool:
    """Return True if the sheet has no data rows (ignoring a single header row)."""
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if not rows:
        return True
    # If every cell in every sampled row is None/empty, treat as empty
    return all(all(v is None or str(v).strip() == "" for v in row) for row in rows)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def expand_one(xlsx_path: Path, dry_run: bool, verbose: bool) -> dict[str, int]:
    """Expand a single xlsx into per-sheet CSVs. Returns action counts."""
    counts = {"new": 0, "updated": 0, "unchanged": 0, "skipped": 0}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"  [ERROR ] {xlsx_path.name}: {exc}", file=sys.stderr)
        counts["skipped"] += 1
        return counts

    out_dir = xlsx_path.parent / xlsx_path.stem
    sheets = wb.sheetnames

    if len(sheets) == 1:
        # Single-sheet xlsx: check if the one sheet has a generic name; still expand
        # so that the ingest tooling can read plain CSV instead of binary xlsx.
        pass

    for sheet_name in sheets:
        ws = wb[sheet_name]

        if is_sheet_empty(ws):
            counts["skipped"] += 1
            if verbose:
                print(f"  [EMPTY ] {xlsx_path.stem}/{sheet_name}  (skipped)")
            continue

        csv_name = sanitize_sheet_name(sheet_name) + ".csv"
        csv_path = out_dir / csv_name
        csv_bytes = sheet_to_csv_bytes(ws)

        if csv_path.exists():
            if sha256_file(csv_path) == sha256_bytes(csv_bytes):
                counts["unchanged"] += 1
                if verbose:
                    print(f"  [SAME  ] {xlsx_path.stem}/{csv_name}")
                continue
            counts["updated"] += 1
            verb = "UPDATE"
        else:
            counts["new"] += 1
            verb = "NEW   "

        print(f"  [{verb}] {xlsx_path.stem}/{csv_name}")
        if not dry_run:
            out_dir.mkdir(parents=True, exist_ok=True)
            csv_path.write_bytes(csv_bytes)

    wb.close()
    return counts


def find_xlsx_files(search_root: Path, recursive: bool) -> list[Path]:
    """Return all .xlsx files under search_root, skipping hidden paths."""
    results: list[Path] = []
    for p in search_root.rglob("*.xlsx") if recursive else search_root.glob("**/*.xlsx"):
        # Skip hidden dirs and the venv
        parts = p.parts
        if any(part.startswith(".") or part in ("venv", ".venv", ".venv-test") for part in parts):
            continue
        # Skip CSVs already produced (they live inside <stem>/ subdirs)
        # — not applicable here since we're only looking for .xlsx
        results.append(p)
    return sorted(results)


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument(
        "--path",
        default=str(DEFAULT_SEARCH_ROOT),
        help=f"Root directory to search for .xlsx files (default: {DEFAULT_SEARCH_ROOT.relative_to(ROOT)})",
    )
    ap.add_argument("--dry-run", action="store_true", help="Report what would be written but do not write")
    ap.add_argument("-v", "--verbose", action="store_true", help="Also print unchanged and skipped sheets")
    ap.add_argument(
        "--recursive", action="store_true", default=True,
        help="Recurse into subdirectories (default: true)",
    )
    args = ap.parse_args()

    search_root = Path(args.path).expanduser().resolve()
    if not search_root.is_dir():
        print(f"ERROR: not a directory: {search_root}", file=sys.stderr)
        return 2

    xlsx_files = find_xlsx_files(search_root, args.recursive)
    if not xlsx_files:
        print(f"No .xlsx files found under {search_root}")
        return 0

    print(f"Found {len(xlsx_files)} .xlsx file(s) under {search_root.relative_to(ROOT)}")
    if args.dry_run:
        print("Mode: DRY RUN — no files will be written")
    print()

    totals = {"new": 0, "updated": 0, "unchanged": 0, "skipped": 0}
    changed_xlsx: list[str] = []

    for xlsx in xlsx_files:
        rel = xlsx.relative_to(ROOT)
        print(f"{rel}")
        counts = expand_one(xlsx, args.dry_run, args.verbose)
        for k in totals:
            totals[k] += counts[k]
        if counts["new"] + counts["updated"] > 0:
            changed_xlsx.append(str(rel))

    print()
    print("=" * 64)
    print("xlsx Expansion Report")
    print("=" * 64)
    print(f"\nTotal: {totals['new']} new | {totals['updated']} updated | "
          f"{totals['unchanged']} unchanged | {totals['skipped']} skipped/empty")

    if changed_xlsx and not args.dry_run:
        print("\nNew/updated CSVs written from:")
        for x in changed_xlsx:
            print(f"  {x}")
        print("\nNext: ingest each CSV directory that changed, e.g.:")
        for x in changed_xlsx:
            stem = Path(x).stem
            parent_slug = Path(x).parent.name
            print(f"  ingest raw/modules/{parent_slug}/{stem}/")

    return 0


if __name__ == "__main__":
    sys.exit(main())
